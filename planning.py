import ffmpeg 
import os
import json
import random
import openpyxl
from openpyxl.styles import PatternFill

os.chdir(r"C:\Users\Lenovo\Desktop\mediaplan")

def hour_to_seconds(time):
    a = 0
    b = ''
    index = 2
    for x in time:
        if ord(x) == ord(':'):
            a += int(b)*pow(60, index)
            index -= 1
            b = ''
            continue
        b += str(x)
    a += int(b)
    return a


def sec_to_hour(time):
    a = int(time/3600)
    b = ''
    if len(str(a)) == 1:
        b = "0" + b + str(a) + ':'
    else:
        b = b + str(a) + ':'
    time -= a*3600 

    a = int(time/60)
    if len(str(a)) == 1:
        b = b + '0' + str(a) + ':'
    else:
        b = b + str(a) + ':'
        
    time -= a*60
    if len(str(time)) == 1:
        b = b + '0' + str(time) 
    else:
        b = b + str(time) 
    return b

def sort_list(repeat):
    pos = 0
    iter = 0
    arr_llst = []
    num = [5, 10, 15, 20]
    for i in range(len(repeat)):
        arr_llst.append(i)
    for j in num:
        for i in repeat[pos:len(repeat)]:
            if i == j:
                if pos == iter:
                    iter += 1
                    pos += 1
                    continue
                b = arr_llst[pos]
                arr_llst[pos] = arr_llst[iter]
                arr_llst[iter] = b
                a = repeat[pos]
                repeat[pos] = repeat[iter]
                repeat[iter] = a
                pos += 1
            iter += 1
        iter = pos
    return arr_llst, repeat

def rearange(index, lists):
    new_list = []
    for i in range(len(index)):
        new_list.append(lists[index[i]])
    return new_list


song_name = []
song_dur = []
object_name = []
object_time1 = []
object_time2 = []
ad_name = []
ad_dur = []
ad_repeat = []

music = r'C:\Users\Lenovo\Desktop\mediaplan\music'
for x in os.listdir(music):
    filename = os.fsdecode(x)
    song_name.append(filename)
    song_dur.append(ffmpeg.probe(music + '\\' + filename)['format']['duration'])

object_file = open(r'C:\Users\Lenovo\Desktop\mediaplan\objects.json', encoding='utf-8')
object_data = json.load(object_file)
for x in object_data['objects']:
    object_name.append(x['Name'])
    object_time1.append(hour_to_seconds(x['time1']))
    object_time2.append(hour_to_seconds(x['time2']))

ad_number = 14
ad_count = 0
ad = r'C:\Users\Lenovo\Desktop\mediaplan\ad'
for x in os.listdir(ad):
    filename = os.fsdecode(x)
    ad_name.append(filename)
    #ad_dur.append(ffmpeg.probe(ad + '\\' + filename)['format']['duration'])
    #ad_dur.append(26)
    #ad_repeat.append(20)
    ad_count += 1
    if ad_count == ad_number:
        break

all_dur = 30 
ad_dur = [30,26,39,24,29,36,33,20,36,34,22,28,31,34,29,26,27,19,35,25,31,32,26,28,30,25,29,27,34,36,37,34,38] 
#all_dur = 19 
#ad_dur = [22,19,16,20,25,16,17,21,20,21,18,21,18,15,22,18,19,20,15,17,17,25,19,21,18,23,15,20,19,15,16,19,20] 
#all_dur = 26 
#ad_dur = [28,24,21,22,19,31,27,23,29,24,25,26,33,28,29,24,21,29,29,28,25,21,29,32,19,26,28,24,26,27,25,26,30]
#all_dur = 37 
#ad_dur = [38,35,37,31,42,34,33,46,41,35,36,34,41,47,31,37,46,34,32,30,39,45,31,44,39,37,36,31,35,39,42,29,35] 

ad_repeat = [15, 20, 20, 20, 10, 20, 15, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 15, 5, 20, 15, 20, 20, 5, 20, 20, 20, 20, 20, 15, 5, 20]

ad_repeat = ad_repeat[0:ad_number]

indeces, ad_repeat = sort_list(ad_repeat)
ad_name = rearange(indeces, ad_name)
ad_dur = rearange(indeces, ad_dur)

print(ad_repeat)

wb = openpyxl.Workbook()

for times in range(len(object_name)):
    index_20_start = len(ad_name)
    for i in range(len(ad_name)):
        if ad_repeat[i] == 20:
            index_20_start = i
            break
    index_20_end = len(ad_name)
    index_20 = index_20_start


    if object_time2[times] < object_time1[times]:
        working_time = 86400 + object_time2[times] - object_time1[times]
    else:
        working_time = object_time2[times] - object_time1[times]
    block_period = 0
    if (len(ad_name)%4 == 0):
        block_period = len(ad_name)/4
    else:
        block_period = int(len(ad_name)/4) + 1

    blocks = block_period*20

    ad_uses = []
    ad_broadcast = []
    for i in range(len(ad_repeat)):
        ad_uses.append(0)
        ad_broadcast.append(0)
    period = int((working_time-20)/blocks)
    cur_time = object_time1[times]
    ad_start = object_time1[times] + period - 120
    cur_block = False
    ad_block = 1
    ads_in_block = 0
    ad_limit = 4
    timer = 0
    timestamps = []
    block_start = ''
    end_time = 0
    if object_time2[times] < object_time1[times]:
        end_time = 86400 + object_time2[times]
    else:
        end_time = object_time2[times] 

    wb.create_sheet(str(int(working_time/3600)))
    ws = wb[str(int(working_time/3600))]
    #with open("Медиаплан" + str(times+1) + ".csv", 'r') as f:
    ws["A1"] = "Имя"
    ws['C1'] = "Время"
    ws["D1"] = str(int(working_time/3600)) + ' часа'
    ws["G1"] = 'Реклама'
    ws["H1"] = len(ad_name)

    ad_sum = 0
    for k in range(len(ad_name)):
        ad_sum += ad_dur[k]*ad_repeat[k]
    percent = ad_sum/working_time

    ws["G2"] = "Повторы"
    ws["H2"] = "20:15:10:5"
    ws["G3"] = 'Продолжительность'
    ws["H3"] = all_dur
    ws["G4"] = "Загруженность"
    ws["H4"] = str(percent*100) + '%'

    row_excel = 2
    colors = []
    green = PatternFill(patternType='solid',
                       fgColor='78D542')
    colors.append(green)
    yellow = PatternFill(patternType='solid',
                       fgColor='FFC638')
    colors.append(yellow)
    cur_color = 1

    while cur_time <end_time:
        block_start = cur_time
        while cur_block == False:
            cur_color = 1
            if cur_time > ad_start:
                #print(cur_time)
                if int(ad_start) - block_start < 0:
                    col = "C"
                    while ws[col + str(row_excel-1)].value is not None:
                        col = chr(ord(col)+1)
                    ws[col + str(row_excel-1)] = "00:00:00"
                    ws[chr(ord(col)+1) + str(row_excel-1)] = "Музыка"
                else:
                    cur_time = ad_start
                    col = "C"
                    while ws[col + str(row_excel-1)].value is not None:
                        col = chr(ord(col)+1)
                    ws[col + str(row_excel-1)] = sec_to_hour(int(cur_time) - block_start)
                    ws[chr(ord(col)+1) + str(row_excel-1)] = "Музыка"
                ad_start += period 
                cur_block = True
                continue
            rand_int = random.randint(0, len(song_name) - 1)
            ws["A" + str(row_excel)] = song_name[rand_int]
            ws["B" + str(row_excel)] = sec_to_hour(cur_time)

            ws['A' + str(row_excel)].fill = colors[cur_color]
            ws['B' + str(row_excel)].fill = colors[cur_color]
            row_excel += 1
            cur_time += int(float(song_dur[rand_int]))+1 

        while cur_block == True:
            cur_color = 0
            if (ad_block-1) == blocks:
                cur_block = False
                continue
            timer = 0
            block_start = cur_time
            if (ad_block)%block_period == 1:
                index_20 = index_20_start
            for i in range(len(ad_repeat)):
                if ad_repeat[i] == 5 and ad_uses[i] != ad_repeat[i]:
                    if (ad_block%int(blocks/5) == 1) or (ad_broadcast[i] == 1):
                        if ads_in_block == 4:
                            ad_broadcast[i] = 1
                            continue
                        ad_broadcast[i] = 0
                        #print(ad_name[i] + " " + str(ad_uses[i] + 1) + " ad block " + str(ad_block)) 
                        ws["A" + str(row_excel)] = ad_name[i]
                        ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                        ws['A' + str(row_excel)].fill = colors[cur_color]
                        ws['B' + str(row_excel)].fill = colors[cur_color]
                        row_excel += 1
                        ads_in_block += 1
                        ad_uses[i] += 1
                        timer += int(float(ad_dur[i])) +1
                        cur_time += int(float(ad_dur[i])) +1 
                        continue

                if ad_repeat[i] == 10 and ad_uses[i] != ad_repeat[i]:
                    if (ad_block%int(blocks/10) == 1) or (ad_broadcast[i] == 1):
                        if ads_in_block == 4:
                            ad_broadcast[i] = 1
                            continue
                        ad_broadcast[i] = 0
                        #print(ad_name[i] + " " + str(ad_uses[i] + 1) + " ad block " + str(ad_block)) 
                        ws["A" + str(row_excel)] = ad_name[i]
                        ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                        ws['A' + str(row_excel)].fill = colors[cur_color]
                        ws['B' + str(row_excel)].fill = colors[cur_color]
                        row_excel += 1
                        ads_in_block += 1
                        ad_uses[i] += 1
                        timer += int(float(ad_dur[i])) +1
                        cur_time += int(float(ad_dur[i])) +1 
                        continue

                if ad_repeat[i] == 15 and ad_uses[i] != ad_repeat[i]:
                    if (ad_block%int(blocks/15) == 1) or (ad_broadcast[i] == 1):
                        if ads_in_block == 4:
                            ad_broadcast[i] = 1
                            continue
                        ad_broadcast[i] = 0
                        #print(ad_name[i] + " " + str(ad_uses[i] + 1) + " ad block " + str(ad_block)) 
                        ws["A" + str(row_excel)] = ad_name[i]
                        ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                        ws['A' + str(row_excel)].fill = colors[cur_color]
                        ws['B' + str(row_excel)].fill = colors[cur_color]
                        row_excel += 1
                        ads_in_block += 1
                        ad_uses[i] += 1
                        timer += int(float(ad_dur[i])) +1
                        cur_time += int(float(ad_dur[i])) +1 
                        continue

                if ad_repeat[i] == 20 and ad_uses[i] != ad_repeat[i]:
                    if (ads_in_block == 4 or i != index_20): 
                        continue
                    #print(ad_name[i] + " " + str(ad_uses[i] + 1) + " ad block " + str(ad_block)) 
                    ws["A" + str(row_excel)] = ad_name[i]
                    ws["B" + str(row_excel)] = sec_to_hour(cur_time)
                    ws['A' + str(row_excel)].fill = colors[cur_color]
                    ws['B' + str(row_excel)].fill = colors[cur_color]
                    row_excel += 1
                    ads_in_block += 1
                    ad_uses[i] += 1
                    index_20 += 1
                    timer += int(float(ad_dur[i])) +1
                    cur_time += int(float(ad_dur[i])) +1 
                    continue
                
            col = "C"
            while ws[col + str(row_excel-1)].value is not None:
                col = chr(ord(col)+1)
            ws[col + str(row_excel-1)] = sec_to_hour(int(cur_time) - block_start)
            ws[chr(ord(col)+1) + str(row_excel-1)] = "Реклама"
            ads_in_block = 0
            ad_block += 1
            cur_block = False
 
    #with open("Медиаплан" + str(times+1) + ".csv", "w", newline='') as f:
    #    write = csv.writer(f, dialect='excel', quoting=csv.QUOTE_ALL, delimiter=",")
    #    write.writerow(fields)
    #    write.writerows(timestamps)
    #    write.writerow([''])
    #    write.writerow([''])
    #    write.writerow(['Повторов за день'])
    #    write.writerows(zip(ad_uses, ad_name))

    ws.append([''])
    ws.append([''])
    ws.append(['Повторов за день'])
    for row in zip(ad_uses, ad_name):
        ws.append(row)
    #    for row in csv.reader(f):
    #        ws.append(row)

    ws.column_dimensions['A'].width = 66
    #os.remove("Медиаплан" + str(times+1) + ".csv")

del wb['Sheet']
wb.save("Медиаплан " + str(ad_number) + " реклам 20,15,10,5 повторов " + str(all_dur) + " сек.xlsx")